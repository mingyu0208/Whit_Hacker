import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = "설명"
worksheet["B1"] = "URL링크"

url = "https://www.malware-traffic-analysis.net/2023/index.html"

req = requests.get(url)
soup = BeautifulSoup(req.text, "lxml")

tags = soup.select("#main_content > div.content > ul > li > a.main_menu")
row = 2
for tag in tags:
    tag_text = tag.text
    tag_href = f"https://www.malware-traffic-analysis.net/2023/{tag['href']}"
    worksheet.cell(row = row, column=1, value=tag_text)
    worksheet.cell(row = row, column=2, value=tag_href)
    row = row + 1

workbook.save(f"malware.xlsx")